import os
from datetime import datetime

import numpy as np
import pandas as pd
from nba_api.stats.endpoints import leaguegamefinder
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine

import builders.gamelog_builder as glb
import globals.global_settings as gls
import globals.global_utils as glu
import globals.run_settings as rns

pd.set_option('display.max_colwidth', None)
pd.set_option('display.max_rows', None)
pd.set_option('display.max_columns', None)
pd.set_option('display.width', None)

#########################################################################################################################################
####  CREATE/UPDATE REAL VALUES TO PREVIOUS PREDICTED TARGET GAMES -> '/DATA/SEASON/'
# Generates the Target values and applies it to Prediction Outputs
# SET 'pull_actual_data' attempt to pull lastest games from API for actual values
# SET 'merge_predictions' to merge all ai predictions for current date
#########################################################################################################################################
pull_actual_data = False
merge_predictions = True

def apply_real_values(pull_actual_data=False, merge_predictions=False):
    print('apply_real_values started...')
    if merge_predictions:
        from globals.merge_predictions import merge_all
        print("Running merge_all.")
        merge_all()
        print("Finished merge_all.")
    latest_season_file = f'{gls.SEASON_DATA_DIR}{rns.prediction_season}.csv'
    if os.path.exists(latest_season_file):
        latest_df = pd.read_csv(latest_season_file)
    if pull_actual_data:
        dataframesarray = find_recent_games()
        if len(dataframesarray) > 0:
            latest_df = pd.concat(dataframesarray, ignore_index=True)
    sorted_directory = sorted(os.listdir(gls.TOP_OUTPUT_DIR), reverse=True)
    for item in sorted_directory:
        check_dir_file = f'{gls.TOP_OUTPUT_DIR}{item}/season_file.csv'
        check_dir_file2 = f'{gls.TOP_OUTPUT_DIR}{item}/season_file2.csv'
        if os.path.exists(check_dir_file2):
            os.remove(check_dir_file2)
        if os.path.exists(check_dir_file):
            print(f'Saved check_dir_file -> {check_dir_file}')
            check_df = pd.read_csv(check_dir_file)
            latest_df['GAME_DATE'] = pd.to_datetime(latest_df['GAME_DATE'])
            check_df['GAME_DATE'] = pd.to_datetime(check_df['GAME_DATE'])
            check_df = check_df.set_index(['PLAYER_ID', 'GAME_DATE'])
            latest_df = latest_df.set_index(['PLAYER_ID', 'GAME_DATE'])
            check_df.update(latest_df[['PTS', 'REB', 'AST']])
            check_df = check_df.reset_index()
            latest_df = latest_df.reset_index()
            check_df.to_csv(check_dir_file2, index=False)
            print(f'Saved check_dir_file2 -> {check_dir_file2}')
    print('apply_real_values completed...')

def calculate_probabilities(sub_df, target_name):
    if not sub_df.iloc[-1]['BET_VAL'] > 0:
        return None, None
    bet_val = sub_df.iloc[-1]['BET_VAL']
    last_5_games = sub_df[target_name].iloc[-6:-1].values
    last_10_games = sub_df[target_name].iloc[-11:-1].values
    prob_5 = np.mean(last_5_games > bet_val) if len(last_5_games) >= 5 else 0
    prob_10 = np.mean(last_10_games > bet_val) if len(last_10_games) >= 10 else 0
    return prob_5, prob_10

def calculate_edge(row):
    prob_5 = row['PROB_5']
    bet_odds = row['BET_ODDS']
    if pd.notna(bet_odds) and bet_odds != 0:
        implied_probability = glu.american_odds_to_implied_probability(bet_odds)
        decimal_odds = glu.american_odds_to_decimal(bet_odds)
        stake = 1
        profit_if_win = (decimal_odds - 1) if bet_odds > 0 else (stake / (bet_odds / -100))
        edge = (prob_5 * profit_if_win) - (implied_probability * stake)
        edge_percentage = round(edge * 100, 2)
    else:
        edge_percentage = 0
    return edge_percentage

def apply_odds_values():
    print('apply_odds_values started...')
    odds_file = f'{gls.ODDS_DATA_DIR}/current_odds-{rns.odds_date}.csv'
    if rns.use_database:
        try:
            mys_sv = gls.SPORTSAI_DBSERVER
            mys_db = gls.SPORTSAI_DBNAME
            mys_us = gls.SPORTSAI_DBUSER
            mys_ps = gls.SPORTSAI_DBPASS
            engine = create_engine(f'mysql+mysqlconnector://{mys_us}:{mys_ps}@{mys_sv}')
            query = f"SELECT * FROM {mys_db}.sportsbook_odds WHERE OpenDate LIKE '%{rns.odds_date}%'"
            result_df = pd.read_sql(query, engine)
            print("Odds data loaded from database.")
        except Exception as e:
            print(f"Database connection failed: {e}")
            result_df = pd.DataFrame(columns=['GAME_DATE', 'PLAYER_NAME', 'BET_TYPE', 'OverValue', 'OverOdds', 'DateChecked', 'EventName'])
    else:
        print("Skipping database connection, 'use_database' is False.")
        result_df = pd.DataFrame(columns=['GAME_DATE', 'PLAYER_NAME', 'BET_TYPE', 'OverValue', 'OverOdds', 'DateChecked', 'EventName'])
    if not result_df.empty:
        result_df = result_df.rename(columns={'OpenDate': 'GAME_DATE', 'BetName': 'PLAYER_NAME', 'BetType': 'BET_TYPE'})
        result_df['GAME_DATE'] = pd.to_datetime(result_df['GAME_DATE']).dt.date
        result_df = result_df.sort_values(by=['PLAYER_NAME', 'GAME_DATE', 'BET_TYPE'])
        result_df.to_csv(odds_file, index=False)
    result_df = pd.read_csv(odds_file) if os.path.exists(odds_file) else pd.DataFrame(columns=['GAME_DATE', 'PLAYER_NAME', 'BET_TYPE', 'OverValue', 'OverOdds', 'DateChecked'])
    grouped = result_df.groupby(['GAME_DATE', 'PLAYER_NAME', 'BET_TYPE']) if not result_df.empty else None
    if grouped:
        agg_df = grouped.agg(
            VAL_RANGE=('OverValue', lambda x: '|'.join(x.astype(str))),
            ODDS_RANGE=('OverOdds', lambda x: '|'.join(x.astype(str))),
            CHANGE_RANGE=('DateChecked', lambda x: '|'.join(x.astype(str)))
        ).reset_index()
        result_df = pd.merge(result_df, agg_df, on=['GAME_DATE', 'PLAYER_NAME', 'BET_TYPE'], how='left')
        result_df = result_df.rename(columns={'OverValue': 'BET_VAL', 'OverOdds': 'BET_ODDS', 'BET_TYPE_x': 'BET_TYPE', 'DateChecked': 'LAST_CHECK', 'EventName': 'EVENT_NAME'})
        result_df['VAL_RANGE'] = result_df['VAL_RANGE'].astype(str).str.replace("'", "", regex=False)
        result_df.to_csv(odds_file, index=False)
    odds_df = pd.read_csv(odds_file) if os.path.exists(odds_file) else pd.DataFrame(columns=['GAME_DATE', 'PLAYER_NAME', 'BET_TYPE', 'BET_VAL', 'BET_ODDS', 'VAL_RANGE'])
    if not odds_df.empty:
        odds_df['GAME_DATE'] = pd.to_datetime(odds_df['GAME_DATE']).dt.date
        odds_sorted = odds_df.sort_values(by=['PLAYER_NAME', 'GAME_DATE', 'BET_TYPE'])
        odds_unique = odds_sorted.drop_duplicates(subset=['PLAYER_NAME', 'GAME_DATE', 'BET_TYPE'], keep='last')
        odds_unique = odds_unique.drop(columns=['UnderValue', 'UnderOdds'], errors='ignore')
        odds_unique.to_csv(odds_file, index=False)
    print('Generated Odds...')
    sorted_directory = sorted(os.listdir(gls.TOP_OUTPUT_DIR), reverse=True)
    for item in sorted_directory:
        check_dir_file = f'{gls.TOP_OUTPUT_DIR}{item}/season_file2.csv'
        if os.path.exists(check_dir_file):
            check_df = pd.read_csv(check_dir_file)
            if 'PTS' in check_df.columns:
                TARGET_NAME = 'PTS'
                df_filtered = odds_unique[odds_unique['BET_TYPE'] == 'Points']
            elif 'REB' in check_df.columns:
                TARGET_NAME = 'REB'
                df_filtered = odds_unique[odds_unique['BET_TYPE'] == 'Rebounds']
            elif 'AST' in check_df.columns:
                TARGET_NAME = 'AST'
                df_filtered = odds_unique[odds_unique['BET_TYPE'] == 'Assists']

            check_df['GAME_DATE'] = pd.to_datetime(check_df['GAME_DATE']).dt.date
            total_ondate_target = check_df.loc[check_df['GAME_DATE'].astype(str) == rns.prediction_date, TARGET_NAME].sum()
            check_df = check_df.merge(df_filtered[['PLAYER_NAME', 'GAME_DATE', 'BET_VAL', 'BET_ODDS', 'VAL_RANGE', 'ODDS_RANGE']],how='left', on=['PLAYER_NAME', 'GAME_DATE'])
            check_df.reset_index(drop=True, inplace=True)
            check_df = check_df.drop(['CORRECT', 'UNDER', 'OVER', 'AVG_OVER', 'AVG_UNDER'], axis=1)
            check_df = check_df.rename(columns={f'{TARGET_NAME}_PRED': 'PRED', 'DIFF_ACT': f'DIFF_{TARGET_NAME}'})
            check_df['BET_VAL'] = check_df['BET_VAL'].fillna(0)
            check_df['OVER_ODDS'] = ((check_df['BET_VAL'] > 0) & (check_df['PRED'] > check_df['BET_VAL'])).astype(int)
            check_df['UNDER_ODDS'] = ((check_df['BET_VAL'] > 0) & (check_df['PRED'] < check_df['BET_VAL'])).astype(int)
            check_df['P>10'] = ((check_df['BET_VAL'] > 0) & (check_df['PRED'] > 1.1 * check_df['BET_VAL'])).astype(int)
            check_df['P<10'] = ((check_df['BET_VAL'] > 0) & (check_df['PRED'] < 0.9 * check_df['BET_VAL'])).astype(int)
            check_df['P>20'] = ((check_df['BET_VAL'] > 0) & (check_df['PRED'] > 1.2 * check_df['BET_VAL'])).astype(int)
            check_df['P<20'] = ((check_df['BET_VAL'] > 0) & (check_df['PRED'] < 0.8 * check_df['BET_VAL'])).astype(int)
            check_df['P>30'] = ((check_df['BET_VAL'] > 0) & (check_df['PRED'] > 1.3 * check_df['BET_VAL'])).astype(int)
            check_df['P<30'] = ((check_df['BET_VAL'] > 0) & (check_df['PRED'] < 0.7 * check_df['BET_VAL'])).astype(int)
            check_df['TIGHT_PRED'] = ((check_df['BET_VAL'] > 0) & (check_df['PRED'] < 1.06 * check_df['BET_VAL']) & (check_df['PRED'] > 0.94 * check_df['BET_VAL'])).astype(int)
            check_df['DIFF_ODDS'] = 0.0
            condition = check_df['GAME_DATE'].astype(str) == rns.prediction_date
            check_df.loc[condition, 'DIFF_ODDS'] = abs(check_df.loc[condition, 'PRED'] - check_df.loc[condition, 'BET_VAL'])
            cols_except_f1_f2 = [col for col in check_df.columns if col not in ['DIFF_AVG', f'DIFF_{TARGET_NAME}', 'DIFF_ODDS', 'VAL_RANGE', 'ODDS_RANGE']]
            new_column_order = cols_except_f1_f2 + ['DIFF_AVG', f'DIFF_{TARGET_NAME}', 'DIFF_ODDS', 'VAL_RANGE', 'ODDS_RANGE']
            check_df = check_df[new_column_order]
            check_df['VAL_RANGE'] = check_df['VAL_RANGE'].astype(str).str.replace("'", "", regex=False)
            percentage_diff = abs(check_df[TARGET_NAME] - check_df['PRED']) / (check_df[TARGET_NAME] + 0.0001)
            absolute_diff = np.abs(check_df[TARGET_NAME] - check_df['PRED'])
            check_df['CORRECT_PRED'] = np.where((percentage_diff <= 0.1) | ((check_df[TARGET_NAME] < 10) & (absolute_diff < 1)),1,0)
            check_df['CORRECT_SUM'] = check_df.groupby(['PLAYER_ID'])['CORRECT_PRED'].cumsum()
            check_df['CORRECT_ODDS'] = 0
            if total_ondate_target > 0:
                check_df['CORRECT_ODDS'] = ((((check_df['BET_VAL'] > 0) & (check_df['PRED'] > check_df['BET_VAL']) & (check_df[TARGET_NAME] > check_df['BET_VAL']))
                                             | ((check_df['BET_VAL'] > 0) & (check_df['PRED'] < check_df['BET_VAL']) & (check_df[TARGET_NAME] < check_df['BET_VAL'])))
                                            & (check_df['GAME_DATE'].astype(str) == rns.prediction_date)).astype(int)
            check_df[f'OVER-10_{TARGET_NAME}'] = (check_df['PRED'] > 1.1 * check_df[TARGET_NAME]).astype(int)
            check_df[f'UNDER-10_{TARGET_NAME}'] = (check_df['PRED'] < 0.9 * check_df[TARGET_NAME]).astype(int)
            check_df['OVER-10_AVG'] = (check_df['PRED'] > 1.1 * check_df[f'AVG_{TARGET_NAME}']).astype(int)
            check_df['UNDER-10_AVG'] = (check_df['PRED'] < 0.9 * check_df[f'AVG_{TARGET_NAME}']).astype(int)
            check_df.reset_index(drop=True, inplace=True)
            folderpath = f'{gls.TOP_OUTPUT_DIR}{item}/'
            check_dir_file3 = f'{folderpath}season_file3.csv'
            check_df = check_df.sort_values(by=['GAME_DATE', 'GAME_ID'])
            check_df.to_csv(check_dir_file3, index=False)
            print(f'Saved check_dir_file3 -> {check_dir_file3}')
            columns_to_move = ['START_POSITION','WEEK_PLAYTIME','LAST_GAME_DAYS','DISTANCE','ALTITUDE','IS_HOME','TEAM_OUT','OPP_OUT',
                               'OPP_PLAYER_ID','OPP_PLAYER_NAME',
                               f'OPP_DEF_{TARGET_NAME}1',f'OPP_DEF_{TARGET_NAME}3',f'OPP_DEF_{TARGET_NAME}AVG','OPP_WEEK_PLAYTIME', 'OPP_LAST_GAME_DAYS',
                               'OPP_DEF_PREVRANK','OPP_DEF_RT3RANK','OPP_DEF_AVGRANK']
            columns_to_move = [col for col in columns_to_move if col in check_df.columns]
            other_columns = [col for col in check_df.columns if col not in columns_to_move]
            new_column_order = other_columns + columns_to_move
            check_df = check_df[new_column_order]
            check_df.reset_index(drop=True, inplace=True)
            check_df.to_csv(check_dir_file3, index=False)

            print('Calculating Probabilities')
            df_sorted = check_df.sort_values(by=['PLAYER_ID', 'GAME_DATE'])
            check_df['PROB_5'] = 0.0
            check_df['PROB_10'] = 0.0
            check_df['PROB_5'] = check_df['PROB_5'].astype(float)
            check_df['PROB_10'] = check_df['PROB_10'].astype(float)
            for player_id, group in df_sorted.groupby('PLAYER_ID'):
                prob_5, prob_10 = calculate_probabilities(group, TARGET_NAME)
                if prob_5 is not None and prob_10 is not None:
                    last_index = group.index[-1]
                    check_df.at[last_index, 'PROB_5'] = prob_5
                    check_df.at[last_index, 'PROB_10'] = prob_10
            check_df = check_df.sort_values(by=['GAME_DATE', 'GAME_ID'])
            check_df.to_csv(check_dir_file3, index=False)

            print('Apply Edge')
            check_df['BET_ODDS'] = pd.to_numeric(check_df['BET_ODDS'], errors='coerce')
            check_df['EDGE'] = 0.0
            check_df['EDGE'] = check_df['EDGE'].astype(float)
            check_df['EDGE'] = check_df.apply(calculate_edge, axis=1)
            check_df = check_df.sort_values(by=['GAME_DATE', 'GAME_ID'])
            check_df.to_csv(check_dir_file3, index=False)

            print("moving officials to end...")
            official_columns = [col for col in check_df.columns if col.startswith("OFFICIAL1_") or col.startswith("OFFICIAL2_")]
            other_columns = [col for col in check_df.columns if col not in official_columns]
            check_df = check_df[other_columns + official_columns]
            check_df = check_df.sort_values(by=['GAME_DATE', 'GAME_ID'])
            check_df.to_csv(check_dir_file3, index=False)

            generate_excelfile(check_df, item, TARGET_NAME)
    print('apply_odds_values completed...')

def generate_excelfile(excel_df, foldername, TARGET_NAME):
    print(f'generate_excelfile {foldername}')
    folderpath = f'{gls.TOP_OUTPUT_DIR}'
    filename = foldername.replace(".", "-")
    excelpath = f'{folderpath}{filename}.xlsx'
    if os.path.exists(excelpath):
        os.remove(excelpath)

    writer = pd.ExcelWriter(excelpath, engine='openpyxl')
    excel_df.to_excel(writer, index=False, sheet_name='MAIN')

    workbook = writer.book
    worksheet = writer.sheets['MAIN']
    max_rows = worksheet.max_row + 1
    print(f'Generated Excel file total rows {max_rows}')


    print(f'Setting Ref Headers...')
    for col in range(67, 155):
        col_letter = get_column_letter(col)
        worksheet.column_dimensions[col_letter].width = 5

    light_yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    light_peach_fill = PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid")
    yellow_columns = ['BO', 'BR', 'BU', 'BX', 'CA', 'CD']
    peach_columns = ['CG', 'CJ', 'CM', 'CP', 'CS', 'CV']

    print(f'Setting Ref Columns...')
    for col in yellow_columns:
        for row in range(1, max_rows):
            cell = worksheet[f'{col}{row}']
            cell.fill = light_yellow_fill
    for col in peach_columns:
        for row in range(1, max_rows):
            cell = worksheet[f'{col}{row}']
            cell.fill = light_peach_fill

    print(f'Setting Ref Value Columns...')
    worksheet['ED1'] = 'BELOW AVG'
    worksheet['EE1'] = 'ABOVE AVG'
    worksheet['EF1'] = 'MINIMUM'
    worksheet['EG1'] = 'LOW4VALUE'
    worksheet['EH1'] = 'HIGH4VALUE'
    worksheet['EI1'] = 'LOW12VALUE'
    worksheet['EJ1'] = 'HIGH12VALUE'
    worksheet['ED2'] = 1.1
    worksheet['EE2'] = 1.1
    worksheet['EF2'] = 2.0
    worksheet['EG2'] = 2
    worksheet['EH2'] = 2
    worksheet['EI2'] = 6
    worksheet['EJ2'] = 8

    print(f'Setting Ref Calc Columns...')
    worksheet['CY1'] = 'LOW4'
    worksheet['CZ1'] = 'HIGH4'
    worksheet['DA1'] = 'LOW12'
    worksheet['DB1'] = 'HIGH12'
    worksheet['DC1'] = 'LOW4_VALID'
    worksheet['DD1'] = 'HIGH4_VALID'
    worksheet['DE1'] = 'LOW12_VALID'
    worksheet['DF1'] = 'HIGH12_VALID'
    worksheet['DI1'] = 'LOW4_SUM'
    worksheet['DJ1'] = 'HIGH4_SUM'
    worksheet['DK1'] = 'LOW12_SUM'
    worksheet['DL1'] = 'HIGH12_SUM'
    worksheet['DN1'] = 'LOW4_COUNT'
    worksheet['DO1'] = 'HIGH4_COUNT'
    worksheet['DP1'] = 'LOW12_COUNT'
    worksheet['DQ1'] = 'HIGH12_COUNT'

    worksheet['DI2'] = f'=SUM(DC2:DC{max_rows})'
    worksheet['DJ2'] = f'=SUM(DD2:DD{max_rows})'
    worksheet['DK2'] = f'=SUM(DE2:DE{max_rows})'
    worksheet['DL2'] = f'=SUM(DF2:DF{max_rows})'
    worksheet['DN2'] = f'=COUNTIF(DC2:DC{worksheet.max_row}, 1) + COUNTIF(DC2:DC{worksheet.max_row}, -1)'
    worksheet['DO2'] = f'=COUNTIF(DD2:DD{worksheet.max_row}, 1) + COUNTIF(DD2:DD{worksheet.max_row}, -1)'
    worksheet['DP2'] = f'=COUNTIF(DE2:DE{worksheet.max_row}, 1) + COUNTIF(DE2:DE{worksheet.max_row}, -1)'
    worksheet['DQ2'] = f'=COUNTIF(DF2:DF{worksheet.max_row}, 1) + COUNTIF(DF2:DF{worksheet.max_row}, -1)'

    header_alignment = Alignment(text_rotation=90, horizontal='center', vertical='center')
    cells_to_align = ['CY1', 'CZ1', 'DA1', 'DB1', 'DC1', 'DD1', 'DE1', 'DF1', 'DI1', 'DJ1', 'DK1', 'DL1', 'ED1', 'EE1', 'EF1', 'EG1', 'EH1', 'EI1','EJ1','DN1','DO1','DP1','DQ1']
    for cell_reference in cells_to_align:
        worksheet[cell_reference].alignment = header_alignment

    print(f'Adding Ref Formulas...')
    add_multiple_formulas(worksheet, max_rows)

    print('Setup Sheet Headers...')
    columns_width = {'PLAYER_ID': .1, 'SEASON': .1,'GAME_ID': 2,'GAME_DATE': 10,'PLAYER_NAME': 13,'TEAM_NAME': 5,
                     f'RTZ_{TARGET_NAME}': 5,f'RT3_{TARGET_NAME}': 5,f'RT5_{TARGET_NAME}': 5,f'RT9_{TARGET_NAME}': 5,
                     f'AVG_{TARGET_NAME}': 5,f'PREV_{TARGET_NAME}': 4,f'LAST_{TARGET_NAME}': 4,TARGET_NAME: 5,'PRED': 5,'CONF': 2,
                     'RT_PCT': 4, 'RT3G_PCT': 4, 'RT5G_PCT': 4, 'RT9G_PCT': 4, 'BET_VAL': 5,
                     'PROB_5':4, 'PROB_10':4, 'EDGE':4,
                     'BET_ODDS': .1, 'OVER_ODDS': 2, 'UNDER_ODDS': 2,
                     'P>10': 2, 'P<10': 2, 'P>20': 2, 'P<20': 2, 'P>30': 2, 'P<30': 2, 'TIGHT_PRED': 2,
                     'DIFF_AVG': 4, f'DIFF_{TARGET_NAME}': 4, 'DIFF_ODDS': 4,
                     'VAL_RANGE': 10,'ODDS_RANGE': .1,
                     'CORRECT_PRED': 3.5,'CORRECT_SUM': 3.5,'CORRECT_ODDS': 3.5,
                     f'OVER-10_{TARGET_NAME}': .1,f'UNDER-10_{TARGET_NAME}': .1,'OVER-10_AVG': 2,'UNDER-10_AVG': 2,
                     'START_POSITION': 3.5,
                     'WEEK_PLAYTIME': 4,
                     'LAST_GAME_DAYS': 3,
                     'DISTANCE': 5,'ALTITUDE': 5,'IS_HOME':2,
                     'TEAM_OUT': 3, 'OPP_OUT': 3,
                     'OPP_PLAYER_ID': .1,'OPP_PLAYER_NAME':13,
                     f'OPP_DEF_{TARGET_NAME}1': 4, f'OPP_DEF_{TARGET_NAME}3': 4, f'OPP_DEF_{TARGET_NAME}AVG': 4, 'OPP_WEEK_PLAYTIME': 4, 'OPP_LAST_GAME_DAYS': 3,
                     'OPP_DEF_PREVRANK': 4, 'OPP_DEF_RT3RANK': 4, 'OPP_DEF_AVGRANK': 4}
    for col_name, width in columns_width.items():
        col_letter = get_column_letter(excel_df.columns.get_loc(col_name) + 1)
        worksheet.column_dimensions[col_letter].width = width
    for idx, col_name in enumerate(excel_df.columns, start=1):
        cell = worksheet.cell(row=1, column=idx)
        cell.alignment = header_alignment
    ####################################################################################################################################
    ######################################################## Apply Initial Styles ######################################################
    print('Apply Initial Styles...')
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.freeze_panes = 'A2'
    worksheet.row_dimensions[1].height = 100
    max_row = len(excel_df) + 1
    main_font = Font(name='Liberation Sans', size=10)

    for row in worksheet.iter_rows():
        for cell in row:
            cell.font = main_font
            if cell.row == 1:
                cell.font = Font(bold=True)

    print('SetColor: (CORRECT_PRED)...')
    lime_green_fill = PatternFill(start_color='00FF00', fill_type='solid')
    for row in worksheet.iter_rows(min_row=2, max_col=len(excel_df.columns), max_row=len(excel_df) + 1):
        if row[excel_df.columns.get_loc('CORRECT_PRED')].value == 1:
            for cell in row:
                cell.fill = lime_green_fill

    print(f'SetColor: (AVG_{TARGET_NAME},{TARGET_NAME},PRED,BET_VAL)...')
    yellow_fill = PatternFill(start_color='FFFF00', fill_type='solid')
    for row in worksheet.iter_rows(max_col=len(excel_df.columns), max_row=max_row):
        for cell in row:
            column_name = excel_df.columns[row.index(cell)]
            if column_name in [f'AVG_{TARGET_NAME}', TARGET_NAME, 'PRED', 'BET_VAL']:
                cell.fill = yellow_fill

    print(f'SetColor: (LAST_{TARGET_NAME},PREV_{TARGET_NAME})...')
    multiply_amt = 1.5 if TARGET_NAME == 'PTS' else 2.0
    bluey_fill = PatternFill(start_color='6FCFFF', fill_type='solid')
    last_column_index = excel_df.columns.get_loc(f'LAST_{TARGET_NAME}') + 1
    avg_column_index = excel_df.columns.get_loc(f'AVG_{TARGET_NAME}') + 1
    prev_column_index = excel_df.columns.get_loc(f'PREV_{TARGET_NAME}') + 1
    for row in worksheet.iter_rows(min_row=2, max_col=len(excel_df.columns), max_row=worksheet.max_row):
        last_cell = row[last_column_index - 1]
        avg_cell = row[avg_column_index - 1]
        prev_cell = row[prev_column_index - 1]
        if last_cell.value is not None and avg_cell.value is not None:
            if last_cell.value >= avg_cell.value * multiply_amt:
                last_cell.fill = bluey_fill
        if prev_cell.value is not None and avg_cell.value is not None:
            if prev_cell.value >= avg_cell.value * multiply_amt:
                prev_cell.fill = bluey_fill

    neonpurple_fill = PatternFill(start_color='BA07F2', fill_type='solid')
    purple_fill = PatternFill(start_color='D0C3FF', fill_type='solid')
    orange_fill1 = PatternFill(start_color='FFCB66', fill_type='solid')
    orange_fill2 = PatternFill(start_color='FFB84A', fill_type='solid')
    orange_fill3 = PatternFill(start_color='E8912B', fill_type='solid')
    orange_fill4 = PatternFill(start_color='E87D15', fill_type='solid')
    orange_fill5 = PatternFill(start_color='E85B05', fill_type='solid')
    blue_fill1 = PatternFill(start_color='CCFFF6', fill_type='solid')
    blue_fill2 = PatternFill(start_color='A1F1FF', fill_type='solid')
    blue_fill3 = PatternFill(start_color='6FCFFF', fill_type='solid')
    blue_fill4 = PatternFill(start_color='3D96FF', fill_type='solid')
    blue_fill5 = PatternFill(start_color='166BFF', fill_type='solid')
    pink_fill1 = PatternFill(start_color='FFD3FC', fill_type='solid')
    pink_fill2 = PatternFill(start_color='FFB1F8', fill_type='solid')
    pink_fill3 = PatternFill(start_color='FF8EF2', fill_type='solid')
    pink_fill4 = PatternFill(start_color='FF6BE7', fill_type='solid')
    tan_fill1 = PatternFill(start_color='FFCDAF', fill_type='solid')
    tan_fill2 = PatternFill(start_color='FFA48F', fill_type='solid')
    tan_fill3 = PatternFill(start_color='FF7777', fill_type='solid')
    orange_range = [orange_fill1,orange_fill2,orange_fill3,orange_fill4,orange_fill5]
    blue_range = [blue_fill1,blue_fill2,blue_fill3,blue_fill4,blue_fill5]
    pink_range = [pink_fill1,pink_fill2,pink_fill3,pink_fill4]
    pink_range2 = [pink_fill1,pink_fill2,pink_fill3]
    tan_range = [tan_fill1, tan_fill2, tan_fill3]


    print('SetColor: (P>10-P<30,TIGHT_PRED,OVER/UNDER,START_POSITION)...')
    set_cell_args = ['EQUAL',['P>10', 'P<10', 'P>20', 'P<20', 'P>30', 'P<30', 'TIGHT_PRED', 'OVER-10_AVG', 'UNDER-10_AVG'],[1],[purple_fill]]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)
    set_cell_args = ['EQUAL_CHAR',['START_POSITION'],['C','G','F'],[purple_fill]]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)

    print(f'SetColor: (RT_PCT,RT3G_PCT,RT5G_PCT,RT9G_PCT,CORRECT_SUM,OPP_DEF_{TARGET_NAME}1,OPP_DEF_{TARGET_NAME}3,OPP_DEF_{TARGET_NAME}AVG,DIFF_AVG,DIFF_ODDS)...')
    set_cell_args = ['GREATERTHAN', ['RT_PCT', 'RT3G_PCT', 'RT5G_PCT'], [.14,.22,.31,.38,.44], orange_range]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)
    set_cell_args = ['GREATERTHAN', ['RT9G_PCT'], [.1,.2,.3,.4,.5], orange_range]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)
    set_cell_args = ['GREATERTHAN', ['CORRECT_SUM'], [2,4,6,8,10], orange_range]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)
    set_cell_args = ['GREATERTHAN', [f'OPP_DEF_{TARGET_NAME}1',f'OPP_DEF_{TARGET_NAME}3'], [7,14,21,28,35], orange_range]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)
    set_cell_args = ['GREATERTHAN', [f'OPP_DEF_{TARGET_NAME}AVG'], [4,8,12,16,20], orange_range]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)
    set_cell_args = ['GREATERTHAN', ['DIFF_AVG','DIFF_ODDS'], [1.5,2.5,3.5,4.5,6], orange_range]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)

    set_cell_args = ['GREATERTHAN', ['PROB_5'], [.55,.65,.75,.85,.95], orange_range]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)
    set_cell_args = ['GREATERTHAN', ['PROB_10'], [.5, .59, .65, .72, .80], orange_range]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)
    set_cell_args = ['GREATERTHAN', ['EDGE'], [0,10,20,30,40], orange_range]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)


    print('SetColor: (TEAM_OUT,OPP_OUT,WEEK_PLAYTIME,OPP_WEEK_PLAYTIME)...')
    set_cell_args = ['GREATERTHAN', ['TEAM_OUT', 'OPP_OUT'], [1,3,5,7,9], blue_range]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)
    set_cell_args = ['GREATERTHAN', ['WEEK_PLAYTIME', 'OPP_WEEK_PLAYTIME'], [25,50,75,100,125], blue_range]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)

    print('SetColor: (DISTANCE,ALTITUDE,LAST_GAME_DAYS,OPP_LAST_GAME_DAYS)...')
    set_cell_args = ['GREATERTHAN', ['DISTANCE'], [400,800,1200,1600], pink_range]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)
    set_cell_args = ['GREATERTHAN', ['ALTITUDE'], [2000,4000,5000], pink_range2]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)
    set_cell_args = ['GREATERTHAN', ['LAST_GAME_DAYS','OPP_LAST_GAME_DAYS'], [5,10,20,30], pink_range]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)

    print('SetColor: (DISTANCE,OPP_DEF_PREVRANK,OPP_DEF_RT3RANK,OPP_DEF_AVGRANK)...')
    set_cell_args = ['GREATERTHAN', ['OPP_DEF_PREVRANK'], [3,5,7], tan_range]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)
    set_cell_args = ['GREATERTHAN', ['OPP_DEF_RT3RANK'], [2.4,5.4,7.9], tan_range]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)
    set_cell_args = ['GREATERTHAN', ['OPP_DEF_AVGRANK'], [1.8,3.4,5.1], tan_range]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)

    print('SetColor: (CORRECT_ODDS)...')
    set_cell_args = ['EQUAL',['CORRECT_ODDS'],[1],[neonpurple_fill]]
    worksheet = set_cell_color(excel_df, worksheet, set_cell_args)

    workbook.save(filename=excelpath)
    print(f'Saved Excel Worksheet: ({worksheet.title}) File -> {excelpath}')

def set_cell_color(excel_df, worksheet, set_args):
    col_indices = [excel_df.columns.get_loc(name) + 1 for name in set_args[1]]
    max_col_idx = max(col_indices)
    for row in worksheet.iter_rows(min_row=2, max_col=max_col_idx, max_row=len(excel_df) + 1):
        for col_idx in col_indices:
            if set_args[0] == 'EQUAL_CHAR':
                if row[col_idx - 1].value in set_args[2]:
                    row[col_idx - 1].fill = set_args[3][0]
            else:
                for idx, value_name in enumerate(set_args[2]):
                    if set_args[0] == 'EQUAL' and row[col_idx - 1].value == value_name:
                        row[col_idx - 1].fill = set_args[3][idx]
                    elif set_args[0] == 'LESSTHAN' and row[col_idx - 1].value < value_name:
                        row[col_idx - 1].fill = set_args[3][idx]
                    elif set_args[0] == 'GREATERTHAN' and row[col_idx - 1].value > value_name:
                        row[col_idx - 1].fill = set_args[3][idx]
    return worksheet

def fetch_game_ids(date):
    date_str = date.strftime('%m/%d/%Y')
    gamefinder = leaguegamefinder.LeagueGameFinder(date_from_nullable=date_str, date_to_nullable=date_str, league_id_nullable='00')
    games_df = gamefinder.get_data_frames()[0]
    print(f'games_df =  {games_df}')
    games_df = games_df[~games_df['SEASON_ID'].astype(str).str.contains('320')]
    game_ids = games_df['GAME_ID'].unique()
    return game_ids.tolist()

def find_recent_games():
    print('find_recent_games...')
    oddsgamedate = datetime.strptime(rns.odds_date, '%Y-%m-%d').date()
    game_ids = fetch_game_ids(oddsgamedate)
    print(f'Checked date {oddsgamedate}.  Returned game_ids ---> {game_ids}')
    dataframesarray = []
    if len(game_ids) > 0:
        for game_id in game_ids:
            newdf = glb.create_game_data(game_id=game_id,will_save=False)
            dataframesarray.append(newdf)
        print("Loaded all games")
    return dataframesarray

def add_multiple_formulas(worksheet, max_rows):
    for row in range(2, max_rows):
        worksheet[f'CY{row}'] = (
            f'=IF(AND(K{row}/$ED$2>BO{row}, K{row}/$ED$2>BR{row}, K{row}/$ED$2>BU{row}, K{row}>$EF$2), '
            f'IF(OR(BO{row}=-1, BR{row}=-1, BU{row}=-1), 0, 1), '
            f'IF(OR(BO{row}=-1, BR{row}=-1, BU{row}=-1), 0, IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(AND(K{row}/$ED$2>BX{row}, K{row}/$ED$2>CA{row}, K{row}/$ED$2>CD{row}, K{row}>$EF$2), '
            f'IF(OR(BX{row}=-1, CA{row}=-1, CD{row}=-1), 0, 1), '
            f'IF(OR(BX{row}=-1, CA{row}=-1, CD{row}=-1), 0, IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(AND(K{row}/$ED$2>CG{row}, K{row}/$ED$2>CJ{row}, K{row}/$ED$2>CM{row}, K{row}>$EF$2), '
            f'IF(OR(CG{row}=-1, CJ{row}=-1, CM{row}=-1), 0, 1), '
            f'IF(OR(CG{row}=-1, CJ{row}=-1, CM{row}=-1), 0, IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(AND(K{row}/$ED$2>CP{row}, K{row}/$ED$2>CS{row}, K{row}/$ED$2>CV{row}, K{row}>$EF$2), '
            f'IF(OR(CP{row}=-1, CS{row}=-1, CV{row}=-1), 0, 1), '
            f'IF(OR(CP{row}=-1, CS{row}=-1, CV{row}=-1), 0, IF(K{row}<$EF$2, 0, -1)))'
        )
        worksheet[f'CZ{row}'] = (
            f'=IF(AND(K{row}*$EE$2<BO{row}, K{row}*$EE$2<BR{row}, K{row}*$EE$2<BU{row}, K{row}>$EF$2), '
            f'IF(OR(BO{row}=-1, BR{row}=-1, BU{row}=-1), 0, 1), '
            f'IF(OR(BO{row}=-1, BR{row}=-1, BU{row}=-1), 0, IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(AND(K{row}*$EE$2<BX{row}, K{row}*$EE$2<CA{row}, K{row}*$EE$2<CD{row}, K{row}>$EF$2), '
            f'IF(OR(BX{row}=-1, CA{row}=-1, CD{row}=-1), 0, 1), '
            f'IF(OR(BX{row}=-1, CA{row}=-1, CD{row}=-1), 0, IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(AND(K{row}*$EE$2<CG{row}, K{row}*$EE$2<CJ{row}, K{row}*$EE$2<CM{row}, K{row}>$EF$2), '
            f'IF(OR(CG{row}=-1, CJ{row}=-1, CM{row}=-1), 0, 1), '
            f'IF(OR(CG{row}=-1, CJ{row}=-1, CM{row}=-1), 0, IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(AND(K{row}*$EE$2<CP{row}, K{row}*$EE$2<CS{row}, K{row}*$EE$2<CV{row}, K{row}>$EF$2), '
            f'IF(OR(CP{row}=-1, CS{row}=-1, CV{row}=-1), 0, 1), '
            f'IF(OR(CP{row}=-1, CS{row}=-1, CV{row}=-1), 0, IF(K{row}<$EF$2, 0, -1)))'
        )
        worksheet[f'DA{row}'] = (
            f'=IF(K{row}/$ED$2>BO{row}, IF(BO{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(BO{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(K{row}/$ED$2>BR{row}, IF(BR{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(BR{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(K{row}/$ED$2>BU{row}, IF(BU{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(BU{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(K{row}/$ED$2>BX{row}, IF(BX{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(BX{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(K{row}/$ED$2>CA{row}, IF(CA{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(CA{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(K{row}/$ED$2>CD{row}, IF(CD{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(CD{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(K{row}/$ED$2>CG{row}, IF(CG{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(CG{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(K{row}/$ED$2>CJ{row}, IF(CJ{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(CJ{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(K{row}/$ED$2>CM{row}, IF(CM{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(CM{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(K{row}/$ED$2>CP{row}, IF(CP{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(CP{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))'
        )
        worksheet[f'DB{row}'] = (
            f'=IF(K{row}*$EE$2<BO{row}, IF(BO{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(BO{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(K{row}*$EE$2<BR{row}, IF(BR{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(BR{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(K{row}*$EE$2<BU{row}, IF(BU{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(BU{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(K{row}*$EE$2<BX{row}, IF(BX{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(BX{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(K{row}*$EE$2<CA{row}, IF(CA{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(CA{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(K{row}*$EE$2<CD{row}, IF(CD{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(CD{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(K{row}*$EE$2<CG{row}, IF(CG{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(CG{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(K{row}*$EE$2<CJ{row}, IF(CJ{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(CJ{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(K{row}*$EE$2<CM{row}, IF(CM{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(CM{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))+'
            f'IF(K{row}*$EE$2<CP{row}, IF(CP{row}=-1, 0, IF(K{row}<$EF$2, 0, 1)), IF(CP{row}=-1, 0, '
            f'IF(K{row}<$EF$2, 0, -1)))'
        )
        worksheet[f'DC{row}'] = f'=IF(CY{row}>=$EG$2, IF(M{row}>K{row}, -1, IF(M{row}<$EF$2, 0, 1)), 0)'
        worksheet[f'DD{row}'] = f'=IF(CZ{row}>=$EH$2, IF(M{row}<K{row}, -1, IF(M{row}<$EF$2, 0, 1)), 0)'
        worksheet[f'DE{row}'] = f'=IF(DA{row}>=$EI$2, IF(M{row}>K{row}, -1, IF(M{row}<$EF$2, 0, 1)), 0)'
        worksheet[f'DF{row}'] = f'=IF(DB{row}>=$EJ$2, IF(M{row}<K{row}, -1, IF(M{row}<$EF$2, 0, 1)), 0)'

print('Starting...')
apply_real_values(pull_actual_data, merge_predictions)
apply_odds_values()
print('Finished...')
